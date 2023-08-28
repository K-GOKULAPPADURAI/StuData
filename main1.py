import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from tkinter import PhotoImage
import openpyxl
import os 
import shutil 
class StudentDetailsApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Student Details Entry")
        self.create_main_page()

    def create_main_page(self):
        main_frame = tk.Frame(self.root)
        main_frame.pack()
        background_image = PhotoImage(file= "backgrounds/pic1.ppm")
        # Create a Label widget to display the background image
        background_label = tk.Label(main_frame, image=background_image)
        background_label.place(relwidth=1, relheight=1)
        
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill="both", expand=True)

        labels = [
            "STUDENT BASIC DETAILS", "FAMILY DETAIL", "ACADEMIC DETAIL", 
            "SCHOLARSHIP DETAILS", "CERTIFICATE DETAILS", "BANK DETAILS","SEARCH STUDENT","UPLOAD EXCEL","EXPORT EXCEL"
        ]

        for label_text in labels:
            frame = tk.Frame(self.notebook)
            self.notebook.add(frame, text=label_text)
            
            if label_text == "STUDENT BASIC DETAILS":
                self.create_student_basic_details(frame)
            elif label_text == "FAMILY DETAIL":
                self.create_family_details(frame)
            elif label_text == "ACADEMIC DETAIL":
                self.create_academic_details(frame)
            elif label_text == "SCHOLARSHIP DETAILS":
                self.create_scholarship_details(frame)
            elif label_text == "CERTIFICATE DETAILS":
                self.create_certificate_details(frame)
            elif label_text == "BANK DETAILS":
                self.create_bank_details(frame)
            elif label_text == "SEARCH STUDENT":
                self.create_search_student_window(frame)
            elif label_text == "UPLOAD EXCEL":
                self.create_upload_excel_window(frame)
            elif label_text == "EXPORT EXCEL":
                self.create_export_excel_window(frame)
        self.notebook.select(0)
        next_button = tk.Button(main_frame, text="Next", command=self.move_to_next_tab, font=("Helvetica", 12))
        next_button.pack(side="bottom", padx=10, pady=10, anchor="e")

    def move_to_next_tab(self):
        current_tab = self.notebook.index(self.notebook.select())
        next_tab = (current_tab + 1) % self.notebook.index(tk.END)
        self.notebook.select(next_tab)

            
    def create_student_basic_details(self, window):
        labels = [
            "APPLICATION NO:", "DATE OF APPLICATION:", "STUDENT NAME:", "AADHAR NUMBER:", "DOB:", "GENDER:",
            "DEPARTMENT:", "QUOTA:", "COMMUNITY:", "CASTE:", "RELIGION:", "MOTHER TONGUE:", "BLOOD GROUP:",
            "MEDIUM OF SCHOOL:", "MARITAL STATUS:", "FATHER NAME:", "MOTHER NAME:", "GUARDIAN NAME:",
            "STUDENT CELL NUMBER:","STUDENT PHOTO:"
        ]

        gender_options = ["Male", "Female"]
        department_options = ["CSE", "MECH", "CIVIL", "EEE", "ECE", "AI/DS"]
        religion_options = ["HINDU", "MUSLIM", "CHRISTIAN", "OTHERS"]
        language_options = ["TAMIL", "ENGLISH", "HINDI", "OTHERS"]
        blood_group_options = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]
        school_medium_options = ["TAMIL", "ENGLISH"]
        marital_status_options = ["Single", "Married"]

        self.student_basic_data = {}

        for idx, label_text in enumerate(labels):
            if idx < len(labels) // 2:
                column = 0
                row = idx
            else:
                column = 2
                row = idx - len(labels) // 2

            label = tk.Label(window, text=label_text)
            label.grid(column=column, row=row, padx=10, pady=5, sticky="w")

            if label_text == "GENDER:":
                gender_var = tk.StringVar(value="Male")
                gender_combobox = ttk.Combobox(window, values=gender_options, textvariable=gender_var)
                gender_combobox.grid(column=column + 1, row=row, padx=10, pady=5, sticky="w")
                self.student_basic_data[label_text] = gender_var
            elif label_text == "DEPARTMENT:":
                department_var = tk.StringVar(value="CSE")
                department_combobox = ttk.Combobox(window, values=department_options, textvariable=department_var)
                department_combobox.grid(column=column + 1, row=row, padx=10, pady=5, sticky="w")
                self.student_basic_data[label_text] = department_var
            elif label_text == "RELIGION:":
                religion_var = tk.StringVar(value="HINDU")
                religion_combobox = ttk.Combobox(window, values=religion_options, textvariable=religion_var)
                religion_combobox.grid(column=column + 1, row=row, padx=10, pady=5, sticky="w")
                self.student_basic_data[label_text] = religion_var
            elif label_text == "MOTHER TONGUE:":
                language_var = tk.StringVar(value="TAMIL")
                language_combobox = ttk.Combobox(window, values=language_options, textvariable=language_var)
                language_combobox.grid(column=column + 1, row=row, padx=10, pady=5, sticky="w")
                self.student_basic_data[label_text] = language_var
            elif label_text == "BLOOD GROUP:":
                blood_group_var = tk.StringVar(value="A+")
                blood_group_combobox = ttk.Combobox(window, values=blood_group_options, textvariable=blood_group_var)
                blood_group_combobox.grid(column=column + 1, row=row, padx=10, pady=5, sticky="w")
                self.student_basic_data[label_text] = blood_group_var
            elif label_text == "MEDIUM OF SCHOOL:":
                medium_var = tk.StringVar(value="TAMIL")
                medium_combobox = ttk.Combobox(window, values=school_medium_options, textvariable=medium_var)
                medium_combobox.grid(column=column + 1, row=row, padx=10, pady=5, sticky="w")
                self.student_basic_data[label_text] = medium_var
            elif label_text == "MARITAL STATUS:":
                marital_var = tk.StringVar(value="Single")
                marital_combobox = ttk.Combobox(window, values=marital_status_options, textvariable=marital_var)
                marital_combobox.grid(column=column + 1, row=row, padx=10, pady=5, sticky="w")
                self.student_basic_data[label_text] = marital_var
            elif label_text == "STUDENT PHOTO:":  # Handle student photo input
                #photo_entry = tk.Entry(window)
                #photo_entry.grid(column=column + 1, row=row, padx=10, pady=5, sticky="w")
                upload_photo_button = tk.Button(window, text="Upload Photo", command=self.upload_photo)
                upload_photo_button.grid(column=column + 1, row=row, padx=10, pady=5, sticky="w")

                # Add a button to open a file dialog for selecting a photo
##                photo_button = tk.Button(window, text="Browse", command=self.browse_photo)
##                photo_button.grid(column=column + 2, row=row, padx=10, pady=5, sticky="w")

            else:
                entry = tk.Entry(window)
                entry.grid(column=column + 1, row=row, padx=10, pady=5, sticky="w")
                self.student_basic_data[label_text] = entry
    def upload_photo(self):
        # Get the student's name from the entry field
        student_name = self.student_basic_data["STUDENT NAME:"].get()

        # Ensure a valid student name is provided
        if student_name:
            # Use the student's name as the file name
            file_name = f"{student_name}.jpg"

            # Use filedialog to open a file dialog for photo selection
            file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.jpeg *.jpg *.gif")])

            if file_path:
                # Check if the "photos" directory exists, create it if not
                if not os.path.exists("photos"):
                    os.mkdir("photos")

                # Save the photo with the student's name as the file name in the "photos" directory
                photo_save_path = os.path.join("photos", file_name)
                shutil.copyfile(file_path, photo_save_path)
                
                messagebox.showinfo("Success", "Photo uploaded and saved successfully!")
        else:
            messagebox.showwarning("Warning", "Please enter the student's name before uploading a photo.")


        #next_button = tk.Button(window, text="Next", command=lambda: self.open_child_window("FAMILY DETAIL"))
        #next_button.grid(column=1, row=len(labels) // 2, padx=10, pady=10, sticky="e")
    def create_family_details(self, window):
            labels = ["ANNUAL INCOME:", "PARENT OCCUPATION:","JOB:","PRESENT ADDRESS:", "PINCODE:", 
                      "DISTRICT:", "RESIDENCE:", "PARENT CELL NUMBER:"]
            
            self.family_data = {}
            
            for label_text in labels:
                label = tk.Label(window, text=label_text)
                label.grid(column=0, row=labels.index(label_text), padx=10, pady=5, sticky="w")
                
                if label_text == "PARENT OCCUPATION:":
                    entry = tk.Entry(window)
                    entry.grid(column=1, row=labels.index(label_text), padx=10, pady=5, sticky="w")
                    self.family_data[label_text] = entry
                elif label_text == "JOB:":
                    job_var = tk.StringVar(value="Government Job")
                    job_combobox = ttk.Combobox(window, values=["Government Job", "Private Job"], textvariable=job_var)
                    job_combobox.grid(column=1, row=labels.index(label_text), padx=10, pady=5, sticky="w")
                    self.family_data[label_text] = job_var
                else:
                    entry = tk.Entry(window)
                    entry.grid(column=1, row=labels.index(label_text), padx=10, pady=5, sticky="w")
                    self.family_data[label_text] = entry

         #   next_button = tk.Button(window, text="Next", command=lambda: self.open_child_window("ACADEMIC DETAIL"))
          #  next_button.grid(column=1, row=len(labels) + 1, padx=10, pady=10, sticky="e")
    def create_academic_details(self, window):
        labels = ["School Name:", "Location:", "Board:", "Year of Passing:", "Percentage of Marks:"]
        grades = ["10th", "11th", "12th"]

        self.academic_data = []

        matrix_rows = 2
        matrix_cols = 2

        for row in range(matrix_rows):
            for col in range(matrix_cols):
                grade_idx = row * matrix_cols + col
                if grade_idx < len(grades):
                    label_text = f"{grades[grade_idx]} Grade"
                    
                    frame = tk.Frame(window)
                    frame.grid(row=row, column=col, padx=10, pady=5, sticky="w")

                    label = tk.Label(frame, text=label_text)
                    label.grid(row=0, column=0, padx=(0, 5), sticky="w")

                    academic_section = {}
                    for sub_row in range(len(labels)):
                        sub_label = tk.Label(frame, text=f"{labels[sub_row]}:")
                        sub_label.grid(row=sub_row + 1, column=0, padx=(0, 5), sticky="w")

                        entry = tk.Entry(frame)
                        entry.grid(row=sub_row + 1, column=1, padx=(0, 10), pady=2, sticky="w")
                        academic_section[labels[sub_row]] = entry

                    self.academic_data.append(academic_section)

        emis_label = tk.Label(window, text="EMIS NO:")
        emis_label.grid(row=matrix_rows * matrix_cols, column=0, padx=(10, 5), pady=5, sticky="w")
        self.emis_entry = tk.Entry(window)
        self.emis_entry.grid(row=matrix_rows * matrix_cols, column=1, padx=(0, 10), pady=2, sticky="w")

        physics_label = tk.Label(window, text="Physics Mark:")
        physics_label.grid(row=matrix_rows * matrix_cols + 1, column=0, padx=(10, 5), pady=5, sticky="w")
        self.physics_entry = tk.Entry(window)
        self.physics_entry.grid(row=matrix_rows * matrix_cols + 1, column=1, padx=(0, 10), pady=2, sticky="w")

        chemistry_label = tk.Label(window, text="Chemistry Mark:")
        chemistry_label.grid(row=matrix_rows * matrix_cols + 2, column=0, padx=(10, 5), pady=5, sticky="w")
        self.chemistry_entry = tk.Entry(window)
        self.chemistry_entry.grid(row=matrix_rows * matrix_cols + 2, column=1, padx=(0, 10), pady=2, sticky="w")

        maths_label = tk.Label(window, text="Maths Mark:")
        maths_label.grid(row=matrix_rows * matrix_cols + 3, column=0, padx=(10, 5), pady=5, sticky="w")
        self.maths_entry = tk.Entry(window)
        self.maths_entry.grid(row=matrix_rows * matrix_cols + 3, column=1, padx=(0, 10), pady=2, sticky="w")

        cutoff_label = tk.Label(window, text="Cut Off Mark:")
        cutoff_label.grid(row=matrix_rows * matrix_cols + 4, column=0, padx=(10, 5), pady=5, sticky="w")
        self.cutoff_entry = tk.Entry(window)
        self.cutoff_entry.grid(row=matrix_rows * matrix_cols + 4, column=1, padx=(0, 10), pady=2, sticky="w")

        #next_button = tk.Button(window, text="Next", command=lambda: self.open_child_window("SCHOLARSHIP DETAILS"))
        #next_button.grid(row=matrix_rows * matrix_cols + 5, column=1, padx=10, pady=10, sticky="e")
    def create_scholarship_details(self, window):
        labels = ["BC:", "MBC:", "PMMS:", "NSP:", "7.5:", "MOOVALUR:", "CKCET:", "FG:","JD:"]
        
        self.scholarship_data = {}
        
        for label_text in labels:
            var = tk.StringVar(value="No")
            checkbox = tk.Checkbutton(window, text=label_text, variable=var, onvalue="Yes", offvalue="No")
            checkbox.grid(column=0, row=labels.index(label_text), padx=10, pady=5, sticky="w")
            self.scholarship_data[label_text] = var

        #next_button = tk.Button(window, text="Next", command=lambda: self.open_child_window("CERTIFICATE DETAILS"))
        #next_button.grid(column=1, row=len(labels), padx=10, pady=10, sticky="e")

    
    def create_certificate_details(self, window):
        labels = ["10th Original:", "11th Original:", "12th Original:", "TC:", "FG:", "JD:", "Community:",
                   "Income:", "Aadhar:", "P-Aadhar:", "Family/Smart card:"]

        self.certificate_data = {}
            
        for label_text in labels:
            var = tk.StringVar(value="No")
            checkbox = tk.Checkbutton(window, text=label_text, variable=var, onvalue="Yes", offvalue="No")
            checkbox.grid(column=0, row=labels.index(label_text), padx=10, pady=5, sticky="w")
            self.certificate_data[label_text] = var

        #next_button = tk.Button(window, text="Next", command=lambda: self.open_child_window("BANK DETAILS"))
        #next_button.grid(column=1, row=len(labels), padx=10, pady=10, sticky="e")

    def create_bank_details(self, window):
        bank_labels = ["BANK NAME:", "BANK BRANCH:", "ACCOUNT NO:", "IFSC CODE:", "MIRC CODE:"]
        self.bank_data = {}

        for label_text in bank_labels:
            label = tk.Label(window, text=label_text)
            label.grid(column=0, row=bank_labels.index(label_text), padx=10, pady=5, sticky="w")

            entry = tk.Entry(window)
            entry.grid(column=1, row=bank_labels.index(label_text), padx=10, pady=5, sticky="w")
            self.bank_data[label_text] = entry

        save_button = tk.Button(window, text="Save", command=self.save_data)
        save_button.grid(column=1, row=len(bank_labels), padx=10, pady=10, sticky="e")
    def create_search_student_window(self, window):
        search_label = tk.Label(window, text="Search Student:")
        search_label.pack(pady=10)

        self.search_entry = tk.Entry(window)
        self.search_entry.pack(pady=5)
        self.search_entry.bind("<KeyRelease>", self.search_students)

        self.suggestions_listbox = tk.Listbox(window, height=5)
        self.suggestions_listbox.pack(pady=5)
        self.suggestions_listbox.bind("<<ListboxSelect>>", self.display_selected_student)

        self.student_info_frame = tk.Frame(window)
        self.student_info_frame.pack(pady=10)

        self.student_photo = tk.Label(self.student_info_frame, text="Photo will be displayed here")
        self.student_photo.grid(row=0, column=0, padx=10, pady=10)

        # Make the student info text field larger
        self.student_info_text = tk.Text(self.student_info_frame, width=50, height=15)
        self.student_info_text.grid(row=0, column=1, padx=10, pady=10)

    def display_selected_student(self, event):
        selected_student = self.suggestions_listbox.get(self.suggestions_listbox.curselection())
        student_info = self.get_student_info(selected_student)

        if student_info:
            # Display student info in the text field
            info_text = "\n".join([f"{key}: {value}" for key, value in student_info.items()])
            self.student_info_text.delete("1.0", tk.END)
            self.student_info_text.insert(tk.END, info_text)

            # Display the student's photo if available
            photo_path = os.path.join(self.photo_directory, f"{selected_student}.jpg")
            if os.path.exists(photo_path):
                photo = tk.PhotoImage(file=photo_path)
                self.student_photo.config(image=photo)
                self.student_photo.image = photo
            else:
                self.student_photo.config(image=None, text="Photo not found")
        else:
            self.student_info_text.delete("1.0", tk.END)
            self.student_photo.config(image=None, text="Student not found")

    def search_students(self, event):
        query = self.search_entry.get().lower()
        self.suggestions_listbox.delete(0, tk.END)
        self.student_names = self.load_student_names_from_excel()
        matching_students = [name for name in self.student_names if name.lower().startswith(query)]
        for student in matching_students:
            self.suggestions_listbox.insert(tk.END, student)

    def get_student_info(self, student_name):
        try:
            # Load the Excel workbook
            wb = openpyxl.load_workbook('student_details.xlsx')
            sheet = wb.active

            # Iterate through rows to find the student by name
            for row in sheet.iter_rows(values_only=True):
                if row[2] == student_name:  # Assuming student name is in the third column (index 2)
                    # Populate the student_info dictionary with all details
                    self.student_info = {
                        "Application No": row[0],  # Adjust column indexes as needed
                        "Date of Application": row[1],
                        "Student Name": row[2],
                        "Aadhar Number": row[3],
                        "DOB": row[4],
                        "Gender": row[5],
                        "Department": row[6],
                        "Quota": row[7],
                        "Community": row[8],
                        "Caste": row[9],
                        "Religion": row[10],
                        "Mother Tongue": row[11],
                        "Blood Group": row[12],
                        "Medium of School": row[13],
                        "Marital Status": row[14],
                        "Father Name": row[15],
                        "Mother Name": row[16],
                        "Guardian Name": row[17],
                        "Student Cell Number": row[18],
                        "STUDENT PHOTO": row[19],  # Assuming the photo path is in the 20th column
                        # Add more fields as needed
                    }
                    return self.student_info

            # If no matching student is found
            return None
        except Exception as e:
            # Handle exceptions (e.g., file not found, sheet not found)
            print(f"Error: {e}")
            return None
    def display_selected_student(self, event):
        selected_index = self.suggestions_listbox.curselection()
        if selected_index:
            selected_index = int(selected_index[0])
            selected_student = self.suggestions_listbox.get(selected_index)
            student_info = self.get_student_info(selected_student)  # Implement this function
            self.display_student_details(student_info)

    def display_student_details(self, student_info):
        info_text = "\n".join([f"{key}: {value}" for key, value in student_info.items()])
        self.student_info_text.delete("1.0", tk.END)
        self.student_info_text.insert(tk.END, info_text)
    def load_student_names_from_excel(self):
        student_names = []
        try:
            wb = openpyxl.load_workbook('student_details.xlsx')
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                student_name = row[2]  # Assuming student name is in the third column (index 2)
                if student_name:
                    student_names.append(student_name)
            wb.close()
        except Exception as e:
            # Handle any exceptions that might occur while loading data
            print(f"Error loading student names: {str(e)}")

        return student_names
        # Placeholder: Implement logic to display the student's photo here (load and display the image)
 
    # Placeholder: Implement logic to display the student's photo here (load and display the image)
    def create_upload_excel_window(self, window):
        #window.title("Upload Excel")
        
        # Create and configure GUI elements for uploading Excel file
        upload_label = tk.Label(window, text="Select Excel File:")
        upload_label.pack(pady=10)
        
        self.uploaded_file_path = tk.StringVar()
        
        upload_button = tk.Button(window, text="Upload", command=self.upload_excel)
        upload_button.pack(pady=5)
        
        append_button = tk.Button(window, text="Append to Current Data", command=self.append_to_excel)
        append_button.pack(pady=10)

    def create_export_excel_window(self, window):
        #window.title("Export Excel")
        
        # Create and configure GUI elements for exporting Excel file
        export_label = tk.Label(window, text="Select Criteria:")
        export_label.pack(pady=10)
        
        self.export_criteria_var = tk.StringVar()
        export_entry = tk.Entry(window, textvariable=self.export_criteria_var)
        export_entry.pack(pady=5)
        
        export_button = tk.Button(window, text="Export Data", command=self.export_excel)
        export_button.pack(pady=10)

    # Implement the functionality for these functions below:

    def search_student(self, student_id):
        # Placeholder: Implement the logic to search for a student and display their information
        student_info = f"Student ID: {student_id}\n"  # Replace with actual student data
        self.student_info_text.delete("1.0", tk.END)  # Clear existing text
        self.student_info_text.insert(tk.END, student_info)
        
        # Placeholder: Implement logic to display the student's photo (you need to load and display the image)
        # Example: Load a sample image for testing
        image = Image.open("sample_student_photo.jpg")
        photo = ImageTk.PhotoImage(image)
        self.student_photo.config(image=photo)
        self.student_photo.image = photo  # Keep a reference to avoid garbage collection

    def upload_excel(self):
        # Placeholder: Implement the logic to upload an Excel file
        file_path = tk.filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.uploaded_file_path.set(file_path)

    def append_to_excel(self):
        # Placeholder: Implement the logic to append data from the uploaded Excel file to the current data
        uploaded_file_path = self.uploaded_file_path.get()
        if uploaded_file_path:
            # Load and append data from the uploaded Excel file
            # Example: Append data using openpyxl or pandas library
            pass

    def export_excel(self):
        criteria = self.export_criteria_var.get()
        if criteria:
            # Placeholder: Implement the logic to export data based on the specified criteria
            # Example: Export data using openpyxl or pandas library
            pass

    def save_data(self):
        wb = openpyxl.load_workbook('student_details.xlsx')
        sheet = wb.active

        data = [
            self.student_basic_data,
            self.family_data,
            *self.academic_data,
            self.scholarship_data,
            self.certificate_data,
            self.bank_data
        ]

        row = []
        for section in data:
            if isinstance(section, dict):
                for value in section.values():
                    if isinstance(value, tk.StringVar):
                        row.append(value.get())
                    elif isinstance(value, tk.Entry):
                        row.append(value.get())  # Retrieve the value from the Entry widget
                    else:
                        row.append(value)
            else:
                row.append(section.get())

        sheet.append(row)
        wb.save('student_details.xlsx')
        wb.close()
        messagebox.showinfo("Success", "Data saved successfully!")

if __name__ == "__main__":
    root = tk.Tk()
    app = StudentDetailsApp(root)
    root.mainloop()
